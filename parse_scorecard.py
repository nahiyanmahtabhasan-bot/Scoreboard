"""Parse the Scorecard sheet from the Project Plan Excel file."""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any, BinaryIO, Sequence

from openpyxl import load_workbook

TEAM_MEMBERS = ["Dhaval", "Syed", "Ahamed", "Yaseen", "Vaizhnavi", "Sneha"]

PHASES = [
    "Process Flow",
    "Functionality",
    "DocType Creation",
    "Validation (@ field level)",
    "Configuration",
    "Testing",
    "Push to GIT",
    "Demo",
    "Feedback -> fixes -> testing ->git hub",
    "Production",
]

PHASE_SHORT_LABELS = {
    "Process Flow": "Process Flow",
    "Functionality": "Functionality",
    "DocType Creation": "DocType Creation",
    "Validation (@ field level)": "Validation",
    "Configuration": "Configuration",
    "Testing": "Testing",
    "Push to GIT": "Push to GIT",
    "Demo": "Demo",
    "Feedback -> fixes -> testing ->git hub": "Feedback Cycle",
    "Production": "Production",
}

Rows = Sequence[tuple[Any, ...]]


def _num(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _str(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    text = str(value).strip()
    return text or None


def _cell(rows: Rows, row_idx: int, col_idx: int) -> Any:
    if row_idx >= len(rows):
        return None
    row = rows[row_idx]
    if col_idx >= len(row):
        return None
    return row[col_idx]


def _load_rows(source: str | Path | BinaryIO) -> Rows:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "Scorecard" not in workbook.sheetnames:
            raise ValueError("Scorecard sheet not found in spreadsheet")
        worksheet = workbook["Scorecard"]
        return list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _parse_module_row(rows: Rows, idx: int) -> dict[str, Any]:
    priority = _num(_cell(rows, idx, 1))
    return {
        "priority": int(priority) if priority is not None else None,
        "module": _str(_cell(rows, idx, 2)),
        "overall": _num(_cell(rows, idx, 3)),
        "phases": {phase: _num(_cell(rows, idx, 4 + i)) for i, phase in enumerate(PHASES)},
        "pct_completed": _num(_cell(rows, idx, 14)),
        "total_items": _num(_cell(rows, idx, 15)),
        "completed_items": _num(_cell(rows, idx, 16)),
    }


def _module_status(module: dict[str, Any]) -> str:
    pct = module.get("pct_completed") or 0
    if pct >= 1.0:
        return "completed"
    if pct <= 0:
        return "not-started"
    return "ongoing"


def _find_section_header(rows: Rows, start: int = 0) -> int | None:
    for idx in range(start, len(rows)):
        if _str(_cell(rows, idx, 1)) == "Priority" and _str(_cell(rows, idx, 2)) == "Modules":
            return idx
    return None


def _module_progress_section(rows: Rows) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    header_row = _find_section_header(rows)
    if header_row is None:
        raise ValueError("Module progress section not found on Scorecard sheet")

    modules: list[dict[str, Any]] = []
    summary: dict[str, Any] | None = None

    for idx in range(header_row + 1, len(rows)):
        module_name = _str(_cell(rows, idx, 2))
        if module_name == "Total Core":
            summary = {
                "overall": _num(_cell(rows, idx, 3)),
                "pct_completed": _num(_cell(rows, idx, 14)),
                "total_items": _num(_cell(rows, idx, 15)),
                "completed_items": _num(_cell(rows, idx, 16)),
            }
            break
        if not module_name:
            continue

        row = _parse_module_row(rows, idx)
        if row["priority"] not in (1, 2):
            continue
        modules.append(row)

    return modules, summary


def _team_module_rows(rows: Rows, start: int, end: int, value_key: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for idx in range(start, min(end + 1, len(rows))):
        module = _str(_cell(rows, idx, 2))
        if not module or module == "Total Core":
            continue
        priority = _num(_cell(rows, idx, 1))
        team_values = {member: _num(_cell(rows, idx, 4 + i)) for i, member in enumerate(TEAM_MEMBERS)}
        result.append(
            {
                "priority": int(priority) if priority is not None else None,
                "module": module,
                "overall": _num(_cell(rows, idx, 3)),
                value_key: team_values,
            }
        )
    return result


def _phase_team_rows(rows: Rows, start: int, end: int, value_key: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for idx in range(start, min(end + 1, len(rows))):
        phase = _str(_cell(rows, idx, 2))
        if not phase or phase == "Modules":
            continue
        team_values = {member: _num(_cell(rows, idx, 4 + i)) for i, member in enumerate(TEAM_MEMBERS)}
        result.append(
            {
                "phase": phase,
                "overall": _num(_cell(rows, idx, 3)),
                value_key: team_values,
            }
        )
    return result


def _find_team_sections(rows: Rows) -> tuple[int, int, int, int]:
    headers = [
        idx
        for idx in range(len(rows))
        if _str(_cell(rows, idx, 1)) == "Priority" and _str(_cell(rows, idx, 2)) == "Modules"
    ]
    if len(headers) < 2:
        return 32, 54, 58, 80

    workload_header = headers[1]
    progress_header = headers[2] if len(headers) > 2 else workload_header + 26

    def section_end(header: int) -> int:
        for idx in range(header + 1, len(rows)):
            if _str(_cell(rows, idx, 2)) == "Total Core":
                return idx - 1
        return header + 22

    return (
        workload_header + 1,
        section_end(workload_header),
        progress_header + 1,
        section_end(progress_header),
    )


def _find_phase_sections(rows: Rows) -> tuple[int, int, int, int]:
    totals_start = totals_end = progress_start = progress_end = 0
    phase_headers = 0

    for idx in range(len(rows)):
        if _str(_cell(rows, idx, 2)) == "Modules" and _str(_cell(rows, idx, 3)) == "Overall":
            phase_headers += 1
            if phase_headers == 1:
                totals_start = idx + 1
            elif phase_headers == 2:
                progress_start = idx + 1
                totals_end = idx - 1
                break

    for idx in range(progress_start, len(rows)):
        if not _str(_cell(rows, idx, 2)):
            progress_end = idx - 1
            break
    else:
        progress_end = min(progress_start + 9, len(rows) - 1)

    return totals_start, totals_end, progress_start, progress_end


def parse_scorecard(source: str | Path | BinaryIO, *, source_label: str | None = None) -> dict[str, Any]:
    if isinstance(source, (str, Path)):
        label = source_label or str(Path(source).resolve())
    else:
        label = source_label or "Live spreadsheet"

    rows = _load_rows(source)
    all_modules, summary = _module_progress_section(rows)
    for module in all_modules:
        module["status"] = _module_status(module)
    module_names = {module["module"] for module in all_modules}

    workload_start, workload_end, progress_start, progress_end = _find_team_sections(rows)
    totals_start, totals_end, progress_start_phases, progress_end_phases = _find_phase_sections(rows)

    team_workload = [
        row
        for row in _team_module_rows(rows, workload_start, workload_end, "tasks")
        if row["module"] in module_names
    ]
    team_progress = [
        row
        for row in _team_module_rows(rows, progress_start, progress_end, "progress")
        if row["module"] in module_names
    ]

    return {
        "source_file": label,
        "source_sheet": "Scorecard",
        "summary": summary or {},
        "modules": all_modules,
        "team_workload": team_workload,
        "team_progress": team_progress,
        "phase_totals": _phase_team_rows(rows, totals_start, totals_end, "totals"),
        "phase_progress": _phase_team_rows(rows, progress_start_phases, progress_end_phases, "progress"),
        "team_members": TEAM_MEMBERS,
        "phases": PHASES,
        "phase_labels": PHASE_SHORT_LABELS,
    }
